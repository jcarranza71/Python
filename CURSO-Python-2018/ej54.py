# xls 2 python
# pip install openpyxl

import openpyxl
doc = openpyxl.load_workbook('data/test.xlsx')

for hoja in doc:
	for fila in hoja:
		print('│ ',end='')
		for columna in fila:
			print(columna.coordinate,':',end=' ')
			print(columna.value,end=' │ ')
		print()

print(doc.sheetnames)
hoja1 = doc['Hoja1']
print(hoja1.title)
print(hoja1['C2'].value)
print(hoja1.cell(row=3,column=3).value)

doc2 = openpyxl.Workbook()
doc2=doc
doc2['Hoja1']['B5'].value=18

doc2.save('data/test2.xlsx')

